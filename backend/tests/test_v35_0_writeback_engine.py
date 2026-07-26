from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

import app.engines.job_engine as engine


class FakeClient:
    target_language_code = "zh-vi"
    source_language_code = "zh"
    bilingual_layout = "inline"
    persistent_cache_hits = 0

    def translate_many(self, texts):
        return ["Khoi dong" for _ in texts]


def test_primary_xlsx_publish_does_not_use_replace(tmp_path: Path, monkeypatch) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    wb = Workbook()
    wb.active["A1"] = "启动"
    wb.save(source)
    wb.close()

    def forbidden_replace(*args, **kwargs):
        raise AssertionError("primary write-back must not call os.replace helper")

    monkeypatch.setattr(engine, "_replace_with_retry", forbidden_replace)
    monkeypatch.setattr(engine, "TranslationClient", lambda *args, **kwargs: FakeClient())
    changed = engine._translate_xlsx(
        source,
        output,
        FakeClient(),
        None,
        {"bilingual_layout": "inline"},
    )

    assert changed == 1
    assert output.exists()
    with ZipFile(output, "r") as archive:
        assert archive.testzip() is None
    wb2 = load_workbook(output)
    assert "Khoi dong" in wb2.active["A1"].value
    wb2.close()


def test_inline_mode_skips_multiline_layout(tmp_path: Path, monkeypatch) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    wb = Workbook()
    wb.active["A1"] = "停止"
    wb.save(source)
    wb.close()

    def forbidden_layout(*args, **kwargs):
        raise AssertionError("inline output must not rebuild workbook layout")

    monkeypatch.setattr(engine, "_apply_excel_multiline_layout", forbidden_layout)
    monkeypatch.setattr(engine, "TranslationClient", lambda *args, **kwargs: FakeClient())
    changed = engine._translate_xlsx(
        source,
        output,
        FakeClient(),
        None,
        {"bilingual_layout": "inline"},
    )
    assert changed == 1
