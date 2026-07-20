from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.engines.job_engine import _reconstruct_plc_configuration, _translate_reconstructed_xlsx


class FakeVietnameseClient:
    target_language_code = "zh-vi"
    source_language_code = "zh"

    def translate_many(self, texts):
        return [f"Bản dịch {index + 1}" for index, _ in enumerate(texts)]


@pytest.mark.parametrize("filename", ["st02(2).xlsx", "st03(1).xlsx"])
def test_plc_hmi_exports_are_compact_deduplicated_and_complete(tmp_path: Path, filename: str):
    source = Path("/mnt/data") / filename
    if not source.exists():
        pytest.skip(f"fixture is not available: {source}")

    cleaned = tmp_path / "cleaned.xlsx"
    translated = tmp_path / "translated.xlsx"
    stats = _reconstruct_plc_configuration(source, cleaned)

    assert stats is not None
    assert stats["mode"] == "enterprise_reconstruction"
    assert stats["summary_counts"]["PLC输入信号"] > 0
    assert stats["summary_counts"]["设备清单"] > 0
    assert stats["summary_counts"]["工位结构"] > 0
    assert stats["summary_counts"]["操作提示"] > 0

    changed = _translate_reconstructed_xlsx(cleaned, translated, FakeVietnameseClient(), None)
    assert changed and changed > 0

    workbook = load_workbook(translated, read_only=True, data_only=False)
    try:
        assert workbook.sheetnames == ["文档概览", "PLC输入信号", "设备清单", "气缸IO配置", "工位结构", "操作提示"]
        for worksheet in workbook.worksheets[1:]:
            values = [
                str(value)
                for row in worksheet.iter_rows(values_only=True)
                for value in row
                if value not in (None, "")
            ]
            assert not any("备用" in value or "Dự phòng" in value for value in values)

        plc = workbook["PLC输入信号"]
        addresses = [str(plc.cell(row, 2).value or "") for row in range(2, plc.max_row + 1)]
        assert len(addresses) == len(set(addresses))
        assert workbook["工位结构"].max_row > 1
        assert workbook["操作提示"].max_row > 1
    finally:
        workbook.close()
