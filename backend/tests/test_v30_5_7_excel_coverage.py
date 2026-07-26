from pathlib import Path

from openpyxl import Workbook, load_workbook

import app.engines.job_engine as job_engine


class EchoClient:
    source_language_code = "zh"
    target_language_code = "zh-vi"
    bilingual_layout = "inline"


class EchoWorkingClient:
    def __init__(self, *args, **kwargs):
        self.source_language_code = "zh"
        self.target_language_code = "vi"
        self.persistent_cache_hits = 0

    def translate_many(self, texts):
        # Simulate the provider omission seen in the customer's workbook.
        return list(texts)

    def invalidate(self, texts):
        return None


def test_mixed_plc_labels_and_repeated_spares_are_filled(tmp_path: Path, monkeypatch):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "X030 气缸IO输入_原点"
    ws["A2"] = "备用"
    ws["B2"] = "备用"
    ws["A3"] = "皮带线IO_正转"
    ws["A4"] = "#N/A"
    wb.save(source)

    monkeypatch.setattr(job_engine, "TranslationClient", EchoWorkingClient)
    changed = job_engine._translate_xlsx(source, output, EchoClient(), None, {"bilingual_layout": "inline"})

    result = load_workbook(output, data_only=False)
    ws = result.active
    assert changed >= 4
    assert ws["A1"].value.startswith("X030 气缸IO输入_原点 - ")
    assert "Vị trí gốc" in ws["A1"].value
    assert ws["A2"].value == "备用 - Dự phòng"
    assert ws["B2"].value == "备用 - Dự phòng"
    assert "quay thuận" in ws["A3"].value.lower()
    assert ws["A4"].value == "#N/A"
    result.close()


def test_translation_unit_preserves_io_and_underscore_context():
    assert job_engine._extract_chinese_translation_unit("X030 气缸IO输入_原点") == "气缸IO输入_原点"
