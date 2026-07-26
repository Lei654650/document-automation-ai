from pathlib import Path
from openpyxl import Workbook, load_workbook

import app.engines.job_engine as engine


class FakeTranslationClient:
    def __init__(self, source_language='auto', target_language='vi', **kwargs):
        self.source_language_code = source_language
        self.target_language_code = target_language
        self.bilingual_layout = 'inline'
        self.persistent_cache_hits = 0

    def translate_many(self, values):
        mapping = {
            '急停报警': 'Báo động dừng khẩn cấp',
            '安全门打开': 'Cửa an toàn đang mở',
        }
        return [mapping.get(value, 'Bản dịch tự động') for value in values]

    def invalidate(self, values):
        return None


def test_pending_marker_is_not_accepted_as_existing_translation():
    assert not engine._valid_existing_target('Cần xác nhận bản dịch')
    assert not engine._valid_existing_target('Translation pending')
    assert engine._valid_existing_target('Báo động dừng khẩn cấp')


def test_translation_qa_repairs_only_pending_cells(tmp_path, monkeypatch):
    source = tmp_path / 'source.xlsx'
    output = tmp_path / 'output.xlsx'
    wb = Workbook()
    ws = wb.active
    ws['A1'] = '急停报警 - Cần xác nhận bản dịch'
    ws['A2'] = '安全门打开 - Cửa an toàn đang mở'
    wb.save(source)

    monkeypatch.setattr(engine, 'TranslationClient', FakeTranslationClient)
    client = FakeTranslationClient(target_language='zh-vi')
    changed = engine._translate_xlsx(source, output, client, None, {'bilingual_layout': 'inline'})

    result = load_workbook(output, data_only=False)
    try:
        assert result.active['A1'].value == '急停报警 - Báo động dừng khẩn cấp'
        assert result.active['A2'].value == '安全门打开 - Cửa an toàn đang mở'
        assert changed == 1
    finally:
        result.close()
    engine._validate_xlsx_delivery_guard(output)
