from pathlib import Path
from openpyxl import Workbook, load_workbook
from app.engines.job_engine import _clean_xlsx

def test_simple_list_keeps_original_coordinates(tmp_path: Path):
    src=tmp_path/"s.xlsx"; out=tmp_path/"o.xlsx"
    wb=Workbook(); ws=wb.active
    for r in range(1,121):
        ws.cell(r,1,f"工位 {r}"); ws.cell(r,2,f"{r}#")
    wb.save(src)
    stats=_clean_xlsx(src,out)
    result=load_workbook(out); sheet=result.active
    assert stats.get("reflow") is None
    assert sheet["A61"].value == "工位 61"
    assert sheet["F1"].value is None
