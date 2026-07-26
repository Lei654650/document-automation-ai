from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.engines import job_engine


class OuterClient:
    source_language_code = "zh"
    target_language_code = "zh-vi"
    bilingual_layout = "inline"
    persistent_cache_hits = 0


class CountingWorkingClient:
    calls = []

    def __init__(self, *args, **kwargs):
        self.source_language_code = "zh"
        self.target_language_code = "vi"
        self.persistent_cache_hits = 0

    def translate_many(self, texts):
        type(self).calls.append(list(texts))
        return ["Bản dịch AI" for _ in texts]


def test_deterministic_plc_labels_skip_provider(tmp_path: Path, monkeypatch):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "X075 备用"
    ws["A2"] = "CY2 载具顶升"
    ws["A3"] = "启动"
    ws["A4"] = "这是一段普通说明文字"
    wb.save(source)

    CountingWorkingClient.calls = []
    monkeypatch.setattr(job_engine, "TranslationClient", CountingWorkingClient)
    job_engine._translate_xlsx(source, output, OuterClient(), None, {"bilingual_layout": "inline"})

    requested = [item for batch in CountingWorkingClient.calls for item in batch]
    assert "备用" not in requested
    assert "载具顶升" not in requested
    assert "启动" not in requested
    assert "这是一段普通说明文字" in requested

    result = load_workbook(output, data_only=False)
    assert "Dự phòng" in result.active["A1"].value
    assert "Xi lanh nâng đồ gá" in result.active["A2"].value
    assert "Khởi động" in result.active["A3"].value
