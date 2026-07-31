from pathlib import Path
import os
import pytest
from openpyxl import load_workbook
from app.engines.job_engine import _reconstruct_plc_configuration, _translate_reconstructed_xlsx

class FakeClient:
    target_language_code='zh-vi'
    def translate_many(self, texts):
        return [f'Bản dịch {i+1}' for i,_ in enumerate(texts)]

def _fixture(name: str) -> Path:
    candidates = [
        Path(os.getenv("DAI_ACCEPTANCE_DATA", "")) / name,
        Path(__file__).resolve().parents[2] / "samples" / name,
        Path("/mnt/data") / name,
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    pytest.skip(f"Optional customer acceptance fixture is unavailable: {name}")

def test_v213_reconstruction_has_adjacent_bilingual_columns(tmp_path: Path):
    source=_fixture('st07.xlsx')
    out=tmp_path/'reconstructed.xlsx'
    stats=_reconstruct_plc_configuration(source,out)
    assert stats and stats['terminology_version']=='21.3'
    wb=load_workbook(out)
    ws=wb['气缸IO配置']
    headers=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
    assert any('越南语原位动作' in str(x) for x in headers)
    assert any('越南语系统' in str(x) for x in headers)
    assert ws['F2'].value in {'缩回','下降','上升','松开','夹紧'}
    assert ws['G2'].value
    assert '——' not in ''.join(str(c.value or '') for row in ws.iter_rows() for c in row)
    wb.close()

def test_v213_translation_fills_unknown_target_cells(tmp_path: Path):
    source=_fixture('st07.xlsx')
    reconstructed=tmp_path/'reconstructed.xlsx'
    final=tmp_path/'final.xlsx'
    _reconstruct_plc_configuration(source,reconstructed)
    count=_translate_reconstructed_xlsx(reconstructed,final,FakeClient(),None)
    assert count is not None
    wb=load_workbook(final)
    for sheet,pairs in {'工位结构':[(2,3),(5,6),(7,8)],'气缸IO配置':[(2,3),(6,7),(8,9),(12,13)]}.items():
        ws=wb[sheet]
        for src,dst in pairs:
            for r in range(2,ws.max_row+1):
                if ws.cell(r,src).value:
                    assert ws.cell(r,dst).value
    wb.close()
