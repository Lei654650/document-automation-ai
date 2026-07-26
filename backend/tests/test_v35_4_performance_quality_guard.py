import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

import app.engines.job_engine as engine


class FakeClient:
    target_language_code = "zh-vi"
    source_language_code = "zh"
    bilingual_layout = "vertical"
    persistent_cache_hits = 0

    def translate_many(self, texts):
        values = []
        for text in texts:
            if "备用" in text:
                values.append("X100 một bên X100")
            else:
                values.append("X101 Khởi động X101")
        return values


def test_progress_callback_cannot_block_document_worker():
    called = []

    def slow_callback(progress, step, message):
        time.sleep(0.4)
        called.append((progress, step, message))

    started = time.perf_counter()
    engine._update(slow_callback, 77, "quality_repair", "普通遥测")
    elapsed = time.perf_counter() - started
    assert elapsed < 0.1
    time.sleep(0.5)
    assert called


def test_vertical_xlsx_uses_separate_target_column_and_dedupes_codes(tmp_path: Path, monkeypatch):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "X100 备用X100"
    ws["A2"] = "X101 启动X101"
    wb.save(source)
    wb.close()

    monkeypatch.setattr(engine, "TranslationClient", lambda *args, **kwargs: FakeClient())
    engine._translate_xlsx(source, output, FakeClient(), None, {"bilingual_layout": "vertical"})
    result = load_workbook(output)
    ws = result.active
    assert ws["A1"].value == "X100 备用"
    assert ws["B1"].value == "X100 Dự phòng"
    assert ws["A2"].value == "X101 启动"
    assert ws["B2"].value == "X101 Khởi động"
    result.close()
