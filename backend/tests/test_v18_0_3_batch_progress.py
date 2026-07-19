from pathlib import Path

from openpyxl import Workbook

from app.engines.job_engine import run_local_job


def _make_xlsx(path: Path, value: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = value
    ws["B1"] = 123
    wb.save(path)


def test_seven_xlsx_batch_moves_beyond_32_and_completes(tmp_path, monkeypatch):
    monkeypatch.setenv("BATCH_MAX_WORKERS", "3")
    sources = []
    upload_dir = tmp_path / "uploads"
    output_dir = tmp_path / "outputs"
    upload_dir.mkdir()
    for index in range(7):
        path = upload_dir / f"st{index + 1:02d}.xlsx"
        _make_xlsx(path, f"file-{index + 1}")
        sources.append((path.name, str(path)))

    events = []
    order = {
        "order_number": "DA-TEST-BATCH",
        "services": ["conversion"],
        "conversion": {"formats": ["original"]},
    }
    result = run_local_job(order, sources, output_dir, lambda p, s, m: events.append((p, s, m)))

    assert result["state"] == "completed"
    assert result["failure_count"] == 0
    assert len(result["outputs"]) == 7
    progresses = [event[0] for event in events]
    assert progresses == sorted(progresses)
    assert any(32 < value < 85 for value in progresses)
    assert progresses[-1] == 100
    assert any("st01.xlsx" in message or "st02.xlsx" in message for _, _, message in events)
