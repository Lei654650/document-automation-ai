from pathlib import Path
from openpyxl import Workbook, load_workbook

from app.engines.job_engine import _clean_xlsx


def _sample(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'] = 'IO显示行数', 8, 'IO显示列数', 4, 'IO地址显示剔除几位'
    ws['A2'], ws['B2'] = '输入名称', 'X'
    ws['A3'], ws['B3'], ws['C3'] = '名称', 0, 1
    ws['A4'], ws['B4'], ws['C4'] = 'M0', 'X000 启动', 'X040 检测夹爪松位'
    ws['A36'], ws['B36'] = '轴名称', 'AX0|拆卸X轴'
    ws['A37'], ws['B37'] = '气缸名称', 'CY0|夹爪'
    ws['A38'], ws['B38'] = '真空名称', 'VA0|拆卸真空'
    ws['A39'], ws['B39'] = '感应器名称', 'SE0|进料感应'
    ws['A50'], ws['B50'] = '气缸IO输入_原点', 'X030'
    ws['A51'], ws['B51'] = '气缸IO输入_动点', 'X031'
    ws['A52'], ws['B52'] = '气缸动作_原位', '缩回'
    ws['A53'], ws['B53'] = '气缸动作_动位', '伸出'
    ws['A54'], ws['B54'] = '气缸IO输出_原点', 'Y030'
    ws['A55'], ws['B55'] = '气缸IO输出_动点', 'Y031'
    ws['A68'], ws['B68'] = 'Tip0_Title', '不在手动模式'
    ws['A86'], ws['B86'] = '工位名称', '工位'
    ws['A87'], ws['B87'] = '工位0', '载具输送'
    ws['A88'], ws['B88'] = '工位0-0', '上层线体'
    wb.save(path)


def test_plc_matrix_is_reconstructed(tmp_path: Path):
    source = tmp_path / 'source.xlsx'
    output = tmp_path / 'output.xlsx'
    _sample(source)
    stats = _clean_xlsx(source, output)
    assert stats['mode'] == 'enterprise_reconstruction'
    wb = load_workbook(output, data_only=False)
    assert wb.sheetnames == ['文档概览', 'PLC输入信号', '设备清单', '气缸IO配置', '工位结构', '操作提示']
    plc = wb['PLC输入信号']
    assert plc['B2'].value == 'X000'
    assert plc['C2'].value == '启动'
    assert plc['E2'].value == '操作按钮'
    stations = wb['工位结构']
    assert stations.max_row == 2
    wb.close()
