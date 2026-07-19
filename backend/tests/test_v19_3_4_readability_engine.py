from pathlib import Path
from openpyxl import Workbook, load_workbook
from app.engines.job_engine import _clean_xlsx


def test_layout_preserves_coordinates_and_uses_readable_dimensions(tmp_path: Path):
    src=tmp_path/'src.xlsx'; out=tmp_path/'out.xlsx'
    wb=Workbook(); ws=wb.active
    ws['A1']='感应器名称\nTên cảm biến'
    ws['B2']='SD0\n前上左安全门\nCửa an toàn trái trước trên'
    ws['M120']='结束'
    wb.save(src)
    stats=_clean_xlsx(src,out)
    result=load_workbook(out); sheet=result.active
    assert sheet['A1'].value=='感应器名称\nTên cảm biến'
    assert sheet['B2'].value.startswith('SD0\n')
    assert sheet['M120'].value=='结束'
    assert sheet.page_setup.fitToWidth == 0
    assert sheet.page_setup.scale == 100
    assert sheet.row_dimensions[2].height >= 50
    assert 8 <= sheet.column_dimensions['B'].width <= 24
    assert stats['structure_changed'] is False
