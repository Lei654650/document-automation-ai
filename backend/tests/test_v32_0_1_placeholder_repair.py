from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.engines.job_engine import _translate_xlsx, _validate_xlsx_delivery_guard
from app.engines.translation_engine import (
    _canonicalize_technical_placeholders,
    _technical_placeholders,
)


class FakeClient:
    target_language_code = "zh-vi"
    source_language_code = "auto"
    bilingual_layout = "inline"
    persistent_cache_hits = 0

    def translate_many(self, texts):
        # The important assertion is that the repair happens before translation.
        assert all("%(WATCH1) - d" not in text for text in texts)
        return ["Trục không ở vị trí %(WATCH1)d, hãy di chuyển đến vị trí %(WATCH1)d." for _ in texts]


def test_canonicalizes_broken_named_printf_placeholder():
    source = "AX1 不在位置%(WATCH1)d，请操作到位置%(WATCH1) - d。"
    fixed = _canonicalize_technical_placeholders(source)
    assert fixed == "AX1 不在位置%(WATCH1)d，请操作到位置%(WATCH1)d。"
    assert _technical_placeholders(fixed) == ["AX1", "%(WATCH1)d", "%(WATCH1)d"]


def test_xlsx_repairs_existing_broken_placeholder_before_delivery(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "AX1 不在位置%(WATCH1)d，请操作到位置%(WATCH1) - d。"
    wb.save(source)

    _translate_xlsx(source, output, FakeClient(), None, {"bilingual_layout": "inline"})
    value = load_workbook(output, data_only=False).active["A1"].value
    assert "%(WATCH1) - d" not in value
    assert value.count("%(WATCH1)d") >= 2
    assert _validate_xlsx_delivery_guard(output)["malformed_placeholders"] == 0
