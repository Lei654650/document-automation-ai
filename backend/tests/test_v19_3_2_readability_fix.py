from pathlib import Path
from openpyxl import Workbook, load_workbook
from app.engines.job_engine import _clean_xlsx

def test_complex_matrix_is_not_moved_or_shrunk(tmp_path: Path):
    src=tmp_path/"m.xlsx"; out=tmp_path/"o.xlsx"
    wb=Workbook(); ws=wb.active
    for c in range(1,41):
        for r in range(1,90): ws.cell(r,c,f"{c}-{r}")
    wb.save(src)
    _clean_xlsx(src,out)
    result=load_workbook(out); sheet=result.active
    assert sheet.max_column == 40
    assert sheet["AN89"].value == "40-89"
    assert sheet.page_setup.fitToWidth == 0
    assert sheet.page_setup.scale == 100

def test_long_list_remains_vertical_and_readable(tmp_path: Path):
    src=tmp_path/"l.xlsx"; out=tmp_path/"o.xlsx"
    wb=Workbook(); ws=wb.active
    for r in range(1,121): ws.cell(r,1,f"工位 {r}")
    wb.save(src)
    _clean_xlsx(src,out)
    result=load_workbook(out); sheet=result.active
    assert sheet["A120"].value == "工位 120"
    assert sheet["F1"].value is None
    assert sheet.row_dimensions[1].height >= 20
