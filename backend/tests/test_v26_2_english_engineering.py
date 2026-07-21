from pathlib import Path
from openpyxl import load_workbook
from app.engines.job_engine import _translate_reconstructed_xlsx, _automation_en_fallback, _polish_automation_en

class EnglishClient:
    target_language_code='zh-en'
    source_language_code='zh'
    def translate_many(self, texts):
        return [f'English translation {i+1}' for i,_ in enumerate(texts)]
    def invalidate(self, texts):
        return None

def test_english_fallback_uses_engineering_terms():
    assert _automation_en_fallback('载具右出料阻挡降位') == 'Pallet Right Outfeed Stopper Lowered Position'
    assert 'Unloading' in _automation_en_fallback('小弹簧下料下层感应')
    assert _polish_automation_en('大弹簧出盘气缸回位','Large spring discharge cylinder return') == 'Large Spring Tray Outfeed Cylinder Retracted Position'

def test_reconstructed_english_overwrites_vietnamese_columns(tmp_path):
    source=Path('/mnt/data/st08_zh-en.xlsx')
    output=tmp_path/'fixed.xlsx'
    count=_translate_reconstructed_xlsx(source,output,EnglishClient(),None)
    assert count and count > 0
    wb=load_workbook(output,read_only=True,data_only=False)
    try:
        ws=wb['PLC输入信号']
        assert ws.cell(1,4).value == 'English Function'
        assert ws.cell(2,4).value == 'Start'
        assert ws.cell(3,4).value == 'Pallet Right Outfeed Stopper Lowered Position'
        assert 'Việt' not in ' '.join(str(c.value or '') for c in ws[1])
        assert ws.cell(2,6).value == 'Operator Controls'
        assert ws.cell(2,8).value == 'Operation and Alarm'
    finally:
        wb.close()
