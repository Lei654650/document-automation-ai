from pathlib import Path
from copy import copy

from openpyxl import Workbook, load_workbook

from app.engines.job_engine import _translate_xlsx


class FakeBatchClient:
    target_language_code = "zh"

    def __init__(self):
        self.request_count = 0

    def translate_many(self, texts):
        self.request_count += 1
        return [f"译:{text}" for text in texts]


def test_excel_translation_uses_batches_and_preserves_non_text(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("EXCEL_TRANSLATION_BATCH_ITEMS", "20")
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "translated.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Pressure setting"
    ws["A2"] = "Pressure setting"  # repeated content
    ws["A3"] = 123.45
    ws["A4"] = "=SUM(A3:A3)"
    for row in range(5, 105):
        ws.cell(row, 1, f"Natural language item {row}")
    font = copy(ws["A1"].font)
    font.bold = True
    ws["A1"].font = font
    wb.save(source)

    client = FakeBatchClient()
    events = []
    changed = _translate_xlsx(source, destination, client, lambda p, s, m: events.append((p, s, m)))

    result = load_workbook(destination)
    out = result.active
    assert changed == 102
    assert out["A1"].value == "译:Pressure setting"
    assert out["A2"].value == "译:Pressure setting"
    assert out["A3"].value == 123.45
    assert out["A4"].value == "=SUM(A3:A3)"
    assert out["A1"].font.bold is True
    assert client.request_count <= 6
    assert events[-1][0] == 80
    assert "约剩余" in events[-1][2]


def test_excel_translation_ignores_declared_empty_max_columns(tmp_path: Path):
    source = tmp_path / "wide.xlsx"
    destination = tmp_path / "translated.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello world"
    # A styled empty cell at Excel's last column must not cause iterating all
    # 16,384 columns times every row.
    ws["XFD1000"].number_format = "0.00"
    wb.save(source)

    client = FakeBatchClient()
    changed = _translate_xlsx(source, destination, client, None)
    assert changed == 1
    assert client.request_count == 1
