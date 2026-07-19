from pathlib import Path
from openpyxl import Workbook, load_workbook

from app.engines.job_engine import _clean_xlsx, _translate_xlsx


def _sample(path: Path) -> None:
    wb = Workbook(); ws = wb.active
    ws['A1'], ws['B1'], ws['C1'] = 'IO显示行数', 8, 'IO显示列数'
    ws['A2'], ws['B2'] = '输入名称', 'X'
    ws['A3'], ws['B3'] = '名称', 0
    ws['A4'], ws['B4'], ws['C4'] = 'M0', 'X000 启动', 'X040 检测夹爪松位'
    ws['A36'], ws['B36'] = '轴名称', 'AX0|拆卸X轴'
    ws['A37'], ws['B37'] = '气缸名称', 'CY0|夹爪'
    ws['A38'], ws['B38'] = '真空名称', 'VA0|拆卸真空'
    ws['A39'], ws['B39'] = '感应器名称', 'SE0|进料感应'
    for rr, value in zip(range(50,56), ['X030','X031','缩回','伸出','Y030','Y031']): ws.cell(rr,2).value=value
    ws['A68'], ws['B68'], ws['C68'], ws['D68'] = 'Tip0_Title', '操作条件不满足', '#N/A', '操作条件不满足'
    ws['A86'], ws['B86'] = '工位名称', '工位'
    ws['A87'], ws['B87'] = '工位0', '载具输送'
    ws['A88'], ws['B88'] = '工位0-0', '上层线体'
    ws['A89'], ws['B89'] = '工位6', '6#'
    wb.save(path)


class FakeClient:
    target_language_code='zh-vi'
    source_language_code='zh'
    persistent_cache_hits=0
    def translate_many(self, texts):
        mapping={'启动':'Khởi động','检测夹爪松位':'Phát hiện vị trí nhả kẹp','拆卸X轴':'Trục X tháo lắp','夹爪':'Kẹp','拆卸真空':'Chân không tháo lắp','进料感应':'Cảm biến cấp liệu','载具输送':'Vận chuyển đồ gá','上层线体':'Băng tải tầng trên','操作条件不满足':'Điều kiện thao tác chưa được đáp ứng'}
        return [mapping.get(t, 'Bản dịch '+t) for t in texts]


def test_three_stage_engineering_reconstruction(tmp_path: Path):
    source=tmp_path/'source.xlsx'; clean=tmp_path/'clean.xlsx'; translated=tmp_path/'translated.xlsx'
    _sample(source)
    stats=_clean_xlsx(source,clean)
    assert stats['mode']=='enterprise_reconstruction'
    wb=load_workbook(clean)
    assert wb['PLC输入信号']['C1'].value=='中文功能'
    assert wb['PLC输入信号']['D1'].value=='越南语功能'
    tips=wb['操作提示']
    assert tips.max_row==2
    assert tips['B2'].value=='操作条件不满足'
    assert tips['D2'].value==2
    assert '#N/A' not in [c.value for row in tips.iter_rows() for c in row]
    stations=wb['工位结构']
    assert stations.max_row==2
    assert stations['E2'].value=='上层线体'
    wb.close()

    changed=_translate_xlsx(clean,translated,FakeClient(),None)
    assert changed>0
    wb=load_workbook(translated)
    assert wb['PLC输入信号']['C2'].value=='启动'
    assert wb['PLC输入信号']['D2'].value=='Khởi động'
    assert wb['操作提示']['C2'].value=='Điều kiện thao tác chưa được đáp ứng'
    assert '——' not in str(wb['PLC输入信号']['D2'].value)
    wb.close()
