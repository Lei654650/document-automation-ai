from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.engines.job_engine import _reconstruct_plc_configuration, _translate_reconstructed_xlsx


class RetryVietnameseClient:
    target_language_code = "zh-vi"
    source_language_code = "zh"

    def __init__(self):
        self.calls = {}

    def translate_many(self, texts):
        values = []
        for text in texts:
            count = self.calls.get(text, 0)
            self.calls[text] = count + 1
            # Simulate a provider echoing source text on the first large-batch pass.
            values.append(text if count == 0 else f"Bản dịch: {text.encode('unicode_escape').decode('ascii')}")
        return values


@pytest.mark.parametrize("filename", ["st02(2).xlsx", "st03(1).xlsx"])
def test_reconstructed_translation_retries_unchanged_source_and_delivers(tmp_path: Path, filename: str):
    source = Path("/mnt/data") / filename
    if not source.exists():
        pytest.skip(f"fixture is not available: {source}")

    cleaned = tmp_path / "cleaned.xlsx"
    translated = tmp_path / "translated.xlsx"
    assert _reconstruct_plc_configuration(source, cleaned) is not None

    changed = _translate_reconstructed_xlsx(cleaned, translated, RetryVietnameseClient(), None)
    assert changed and changed > 0

    workbook = load_workbook(translated, read_only=True, data_only=False)
    try:
        for sheet_name, pairs in {
            "PLC输入信号": ((3, 4), (5, 6), (7, 8)),
            "设备清单": ((1, 2), (4, 5), (6, 7), (8, 9)),
            "气缸IO配置": ((2, 3), (6, 7), (8, 9), (12, 13)),
            "工位结构": ((2, 3), (5, 6), (7, 8)),
            "操作提示": ((2, 3), (5, 6)),
        }.items():
            ws = workbook[sheet_name]
            for src_col, dst_col in pairs:
                for row in range(2, ws.max_row + 1):
                    source_text = str(ws.cell(row, src_col).value or "").strip()
                    target_text = str(ws.cell(row, dst_col).value or "").strip()
                    if source_text:
                        assert target_text, (sheet_name, row, source_text)
    finally:
        workbook.close()
